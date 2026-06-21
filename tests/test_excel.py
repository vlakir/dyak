"""Тесты чтения и маппинга xlsx (этап 0, T001)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from dyak.config import Config
from dyak.errors import TableError
from dyak.io.excel import read_table

_HEADERS = ['Фамилия', 'Имя', 'Отчество', 'Должность', 'Пол', 'Дата начала']
_ROWS = [
    ['Иванов', 'Пётр', 'Семёнович', 'директор', 'м', '01.07.2026'],
    ['Петрова', 'Анна', 'Сергеевна', 'главный бухгалтер', 'ж', '01.07.2026'],
]


def _make_config() -> Config:
    return Config.model_validate(
        {
            'columns': {
                'surname': 'Фамилия',
                'name': 'Имя',
                'patronymic': 'Отчество',
                'position': 'Должность',
                'gender': 'Пол',
                'дата_начала': 'Дата начала',
            },
            'filename': '{{ сотрудник.фамилия }}.docx',
        },
    )


def _make_xlsx(
    path: Path,
    headers: list[str],
    rows: list[list[str]],
    sheet_title: str | None = None,
) -> Path:
    wb = Workbook()
    ws = wb.active
    if sheet_title is not None:
        ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_reads_all_rows(tmp_path: Path) -> None:
    xlsx = _make_xlsx(tmp_path / 't.xlsx', _HEADERS, _ROWS)
    people = read_table(xlsx, _make_config())
    assert len(people) == 2
    assert people[0].surname == 'Иванов'
    assert people[0].position == 'директор'
    assert people[1].name == 'Анна'


def test_extra_columns_available(tmp_path: Path) -> None:
    xlsx = _make_xlsx(tmp_path / 't.xlsx', _HEADERS, _ROWS)
    people = read_table(xlsx, _make_config())
    assert people[0].extra['дата_начала'] == '01.07.2026'


def test_missing_required_column_raises(tmp_path: Path) -> None:
    headers = [h for h in _HEADERS if h != 'Пол']
    rows = [[c for c, h in zip(r, _HEADERS, strict=True) if h != 'Пол'] for r in _ROWS]
    xlsx = _make_xlsx(tmp_path / 't.xlsx', headers, rows)
    with pytest.raises(TableError, match='Пол'):
        read_table(xlsx, _make_config())


def test_missing_optional_column_raises(tmp_path: Path) -> None:
    headers = [h for h in _HEADERS if h != 'Дата начала']
    rows = [r[:-1] for r in _ROWS]
    xlsx = _make_xlsx(tmp_path / 't.xlsx', headers, rows)
    with pytest.raises(TableError, match='Дата начала'):
        read_table(xlsx, _make_config())


def test_empty_required_cell_reports_row_number(tmp_path: Path) -> None:
    rows = [['Иванов', 'Пётр', 'Семёнович', 'директор', 'м', '01.07.2026'],
            ['', 'Анна', 'Сергеевна', 'бухгалтер', 'ж', '01.07.2026']]
    xlsx = _make_xlsx(tmp_path / 't.xlsx', _HEADERS, rows)
    with pytest.raises(TableError, match='3'):  # лист: заголовок=1, 2-я данных=строка 3
        read_table(xlsx, _make_config())


def test_sheet_not_found_raises(tmp_path: Path) -> None:
    xlsx = _make_xlsx(tmp_path / 't.xlsx', _HEADERS, _ROWS)
    with pytest.raises(TableError, match='не найден'):
        read_table(xlsx, _make_config(), sheet='Нет такого')


def test_empty_table_raises(tmp_path: Path) -> None:
    path = tmp_path / 't.xlsx'
    Workbook().save(path)
    with pytest.raises(TableError, match='пуст'):
        read_table(path, _make_config())


def test_numeric_cell_rendered_without_decimal(tmp_path: Path) -> None:
    path = tmp_path / 't.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append([*_HEADERS, 'Номер'])
    ws.append([*_ROWS[0], 17])  # целое число хранится как 17, не «17.0»
    wb.save(path)
    config = Config.model_validate(
        {
            'columns': {
                'surname': 'Фамилия',
                'name': 'Имя',
                'patronymic': 'Отчество',
                'position': 'Должность',
                'gender': 'Пол',
                'дата_начала': 'Дата начала',
                'номер': 'Номер',
            },
            'filename': '{{ сотрудник.фамилия }}.docx',
        },
    )
    people = read_table(path, config)
    assert people[0].extra['номер'] == '17'


def test_sheet_selected_by_name(tmp_path: Path) -> None:
    path = tmp_path / 't.xlsx'
    wb = Workbook()
    wb.active.title = 'Прочее'
    target = wb.create_sheet('Сотрудники')
    target.append(_HEADERS)
    for row in _ROWS:
        target.append(row)
    wb.save(path)
    people = read_table(path, _make_config(), sheet='Сотрудники')
    assert len(people) == 2
    assert people[0].surname == 'Иванов'
