"""
Чтение входной таблицы xlsx (T006 + учёт фильтра T019).

Колонки читаются по их заголовкам: заголовок нормализуется (пробелы → `_`)
и становится ключом контекста. Обязательного маппинга больше нет —
стандартные кадровые колонки распознаются по названию (`columns.recognize`),
прочие доступны в шаблоне по нормализованному имени. Значения берутся как
строки ровно как в ячейке (§12): даты/номера, введённые текстом, openpyxl
возвращает без переформатирования.

**Скрытые строки (T019).** Если в таблице включён автофильтр (или строки
скрыты вручную), Excel помечает отфильтрованные строки `hidden`. Документы
формируются **только из видимых** строк, но контентное распознавание ролей
(T016) использует образцы со **всех** строк, включая скрытые, — чтобы фильтр
не обеднял распознавание. Поэтому книга грузится **без** `read_only`: в
read-only режиме openpyxl не отдаёт `row_dimensions`/`hidden` (ADR 2026-06-23).
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from dyak.columns import NAME, PATRONYMIC, SURNAME, normalize_header, recognize
from dyak.domain import Person, Table
from dyak.errors import TableError

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from dyak.config import Config

logger = logging.getLogger(__name__)

# Сколько непустых ячеек колонки брать как образцы для контентного
# распознавания роли (T016). Достаточно небольшой выборки — таблицы
# однородны по колонке.
_SAMPLE_LIMIT = 5


def _to_str(value: object) -> str:
    """Привести значение ячейки к строке без переформатирования."""
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _person_has_name(
    cells: dict[str, str], roles: dict[str, str], fullname_source: str | None
) -> bool:
    """Есть ли в строке хоть какая-то часть ФИО (иначе строка — пустой человек)."""
    if fullname_source is not None:
        return bool(cells.get(fullname_source, '').strip())
    return any(
        cells.get(roles.get(role, ''), '').strip()
        for role in (SURNAME, NAME, PATRONYMIC)
    )


def _drop_nameless_rows(
    people: list[Person], roles: dict[str, str], fullname_source: str | None
) -> list[Person]:
    """
    Отсеять строки без ФИО, если таблица ФИО-ориентированная (T029).

    Частая штатная ситуация — пред-пронумерованные пустые строки («№ п/п» есть,
    остального нет): по ним нечего генерировать (ФИО — основа кадрового
    документа), а имя файла по умолчанию (`{{ Фамилия … }}`) на них падало.
    Если колонок ФИО нет вовсе (генерация по другому ключу) — строки не трогаем.
    """
    has_fio = fullname_source is not None or any(
        role in roles for role in (SURNAME, NAME, PATRONYMIC)
    )
    if not has_fio:
        return people
    kept = [p for p in people if _person_has_name(p.cells, roles, fullname_source)]
    dropped = len(people) - len(kept)
    if dropped:
        logger.warning(
            'Пропущено строк без ФИО: %d — документы по ним не формируются '
            '(пустые/пред-пронумерованные строки)',
            dropped,
        )
    return kept


def _select_sheet(wb: Workbook, sheet: str | None) -> Worksheet:
    if sheet is None:
        ws = wb.active
        if ws is None:
            msg = 'В книге нет активного листа'
            raise TableError(msg)
        return ws
    if sheet not in wb.sheetnames:
        msg = f'Лист не найден: {sheet}'
        raise TableError(msg)
    return wb[sheet]


def _build_columns(header_row: tuple[object, ...]) -> list[tuple[int, str]]:
    """Сопоставить индексам колонок нормализованные ключи; ловить коллизии."""
    columns: list[tuple[int, str]] = []
    seen: dict[str, str] = {}  # нормализованный ключ → исходный заголовок
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        raw_str = str(raw)
        key = normalize_header(raw_str)
        if key in seen and seen[key] != raw_str:
            msg = (
                f'Колонки «{seen[key]}» и «{raw_str}» дают одинаковый ключ '
                f'«{key}» после нормализации пробелов'
            )
            raise TableError(msg)
        seen[key] = raw_str
        columns.append((idx, key))
    if not columns:
        msg = 'В строке заголовков нет ни одной колонки'
        raise TableError(msg)
    return columns


def read_table(path: Path, config: Config, sheet: str | None = None) -> Table:
    """
    Прочитать xlsx в `Table` (роли колонок + видимые строки по заголовкам).

    Грузим **без** `read_only`: иначе openpyxl не отдаёт состояние скрытых
    строк (фильтр), а оно нужно, чтобы не генерировать документы по
    отфильтрованным строкам (T019).
    """
    wb = load_workbook(path, data_only=True)
    try:
        return _read_rows(_select_sheet(wb, sheet), config)
    finally:
        wb.close()


def _read_rows(ws: Worksheet, config: Config) -> Table:
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        msg = 'Таблица пуста: нет строки заголовков'
        raise TableError(msg)

    columns = _build_columns(header_row)

    # Образцы ячеек для контентного распознавания ролей (T016): первые
    # `_SAMPLE_LIMIT` непустых значений каждой колонки. Собираем со ВСЕХ строк
    # (видимых и скрытых, T019) — фильтр не должен обеднять распознавание.
    samples: dict[str, list[str]] = {key: [] for _, key in columns}
    people: list[Person] = []
    hidden_skipped = 0
    # Заголовок — первая строка (`iter_rows` стартует с min_row=1); данные — со
    # строки 2. `sheet_row` нужен для проверки `hidden` через `row_dimensions`.
    for sheet_row, row in enumerate(rows, start=2):
        cells = {
            key: _to_str(row[idx]) if idx < len(row) else '' for idx, key in columns
        }
        if all(value == '' for value in cells.values()):
            continue  # пропускаем полностью пустые строки (хвост листа)
        for _, key in columns:
            bucket = samples[key]
            if cells[key] and len(bucket) < _SAMPLE_LIMIT:
                bucket.append(cells[key])
        if ws.row_dimensions[sheet_row].hidden:
            hidden_skipped += 1  # скрытая фильтром/вручную — не генерируем
            continue
        people.append(Person(cells=cells))

    if hidden_skipped:
        logger.info(
            'Пропущено скрытых (отфильтрованных) строк: %d — документы по ним '
            'не формируются; в распознавании ролей их данные учтены',
            hidden_skipped,
        )

    recognition = recognize([key for _, key in columns], config.aliases, samples)
    people = _drop_nameless_rows(people, recognition.roles, recognition.fullname_source)
    return Table(
        roles=recognition.roles,
        fullname_source=recognition.fullname_source,
        people=people,
    )
